"""知识库加载与混合检索（TF-IDF + 关键词匹配）"""
import csv
import json
import os
import re
from io import BytesIO, StringIO

import jieba
import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

DATA_PATH = os.path.join(os.path.dirname(__file__), "data", "qa_knowledge.json")

_qa_data: list[dict] = []
_vectorizer: TfidfVectorizer | None = None
_tfidf_matrix = None
_question_tokens: list[set[str]] = []

_STOPWORDS = frozenset(
    "的 了 在 是 我 有 和 就 不 人 都 一 一个 上 也 很 到 说 要 去 你 会 着 没有 看 好 自己 这".split()
)

_SUPPORTED_EXTS = {".json", ".xlsx", ".csv"}


def _tokenize(text: str) -> str:
    return " ".join(w for w in jieba.cut(text) if w.strip() and w not in _STOPWORDS)


def _keyword_set(text: str) -> set[str]:
    return {w for w in jieba.cut(text) if len(w) >= 2 and w not in _STOPWORDS}


def _normalize_qa_items(raw_items: list[dict]) -> list[dict]:
    normalized: list[dict] = []
    for item in raw_items:
        question = str(item.get("question", "")).strip()
        answer = str(item.get("answer", "")).strip()
        if not question or not answer:
            continue
        if answer == "答案正在整理中":
            continue
        normalized.append({
            "id": len(normalized) + 1,
            "question": question,
            "answer": answer,
        })
    return normalized


def _parse_json_content(content: bytes) -> list[dict]:
    try:
        payload = json.loads(content.decode("utf-8-sig"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"JSON 解析失败：{exc}") from exc

    if isinstance(payload, dict) and isinstance(payload.get("data"), list):
        payload = payload["data"]
    if not isinstance(payload, list):
        raise ValueError("JSON 文件格式错误：应为数组，元素包含 question / answer 字段。")

    items = [item for item in payload if isinstance(item, dict)]
    return _normalize_qa_items(items)


def _parse_csv_content(content: bytes) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV 编码不支持，请使用 UTF-8 编码。") from exc

    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV 文件缺少表头。")

    fieldnames = [f.strip() for f in reader.fieldnames if f]
    q_key = next((f for f in fieldnames if "问题" in f or f.lower() in {"question", "q"}), "")
    a_key = next((f for f in fieldnames if "答案" in f or f.lower() in {"answer", "a"}), "")
    if not q_key or not a_key:
        raise ValueError("CSV 表头需包含“问题/答案”字段（或 question/answer）。")

    rows: list[dict] = []
    for row in reader:
        rows.append({
            "question": row.get(q_key, ""),
            "answer": row.get(a_key, ""),
        })
    return _normalize_qa_items(rows)


def _parse_xlsx_sheet(ws) -> list[dict]:
    rows: list[dict] = []

    if ws.title == "精选问题":
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if not row:
                continue
            question = row[1] if len(row) > 1 else ""
            answer = row[2] if len(row) > 2 else ""
            rows.append({"question": question, "answer": answer})
        parsed = _normalize_qa_items(rows)
        if parsed:
            return parsed

    header_row_idx = -1
    q_col = -1
    a_col = -1
    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=True),
        start=1,
    ):
        if not row:
            continue
        cells = [str(cell).strip() if cell is not None else "" for cell in row]
        if q_col < 0:
            q_col = next((i for i, c in enumerate(cells) if "问题" in c or c.lower() == "question"), -1)
        if a_col < 0:
            a_col = next((i for i, c in enumerate(cells) if "答案" in c or c.lower() == "answer"), -1)
        if q_col >= 0 and a_col >= 0:
            header_row_idx = idx
            break

    if header_row_idx < 0:
        return []

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
        if not row:
            continue
        question = row[q_col] if len(row) > q_col else ""
        answer = row[a_col] if len(row) > a_col else ""
        rows.append({"question": question, "answer": answer})
    return _normalize_qa_items(rows)


def _parse_xlsx_content(content: bytes) -> list[dict]:
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"Excel 解析失败：{exc}") from exc

    if "精选问题" in wb.sheetnames:
        parsed = _parse_xlsx_sheet(wb["精选问题"])
        if parsed:
            return parsed

    for name in wb.sheetnames:
        parsed = _parse_xlsx_sheet(wb[name])
        if parsed:
            return parsed

    return []


def import_knowledge_from_bytes(filename: str, content: bytes) -> int:
    if not filename:
        raise ValueError("文件名不能为空。")
    if not content:
        raise ValueError("上传文件为空。")

    ext = os.path.splitext(filename)[1].lower()
    if ext not in _SUPPORTED_EXTS:
        raise ValueError("仅支持 .xlsx / .json / .csv 文件。")

    if ext == ".xlsx":
        qa_list = _parse_xlsx_content(content)
    elif ext == ".json":
        qa_list = _parse_json_content(content)
    else:
        qa_list = _parse_csv_content(content)

    if not qa_list:
        raise ValueError("未解析到有效问答，请检查文件内容或模板格式。")

    os.makedirs(os.path.dirname(DATA_PATH), exist_ok=True)
    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(qa_list, f, ensure_ascii=False, indent=2)

    load_knowledge()
    return len(qa_list)


def get_knowledge_stats() -> dict:
    if _vectorizer is None or _tfidf_matrix is None:
        load_knowledge()
    return {"count": len(_qa_data)}


def load_knowledge():
    global _qa_data, _vectorizer, _tfidf_matrix, _question_tokens
    with open(DATA_PATH, "r", encoding="utf-8") as f:
        _qa_data = json.load(f)

    if not _qa_data:
        _vectorizer = TfidfVectorizer()
        _tfidf_matrix = _vectorizer.fit_transform([""])
        _question_tokens = [set()]
        print("知识库加载完成：0 条问答对")
        return

    questions_tokenized = [_tokenize(item["question"]) for item in _qa_data]
    _question_tokens = [_keyword_set(item["question"]) for item in _qa_data]

    _vectorizer = TfidfVectorizer()
    _tfidf_matrix = _vectorizer.fit_transform(questions_tokenized)
    print(f"知识库加载完成：{len(_qa_data)} 条问答对")


def search(query: str, top_k: int = 3) -> list[dict]:
    if _vectorizer is None or _tfidf_matrix is None:
        load_knowledge()

    query_tokens = _keyword_set(query)
    query_tokenized = _tokenize(query)

    query_vec = _vectorizer.transform([query_tokenized])
    tfidf_scores = cosine_similarity(query_vec, _tfidf_matrix).flatten()

    keyword_scores = []
    for qt in _question_tokens:
        if not qt or not query_tokens:
            keyword_scores.append(0.0)
            continue
        overlap = len(query_tokens & qt)
        keyword_scores.append(overlap / max(len(query_tokens), len(qt)))

    combined = []
    for i in range(len(_qa_data)):
        score = 0.7 * tfidf_scores[i] + 0.3 * keyword_scores[i]
        if score > 0:
            combined.append((i, round(score, 4)))

    combined.sort(key=lambda x: x[1], reverse=True)

    results = []
    for idx, score in combined[:top_k]:
        results.append({
            "question": _qa_data[idx]["question"],
            "answer": _qa_data[idx]["answer"],
            "score": score,
        })
    return results
