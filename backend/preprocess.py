"""将 xlsx 知识库数据预处理为 JSON 格式"""
import json
import os
import openpyxl

XLSX_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "data-source",
    "需要准备的数据-吴迪260313.xlsx",
)
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "data", "qa_knowledge.json")


def preprocess():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["精选问题"]

    qa_list = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        seq, question, answer = row[0], row[1], row[2]
        if not question or not answer:
            continue
        if answer.strip() == "答案正在整理中":
            continue
        qa_list.append({
            "id": int(seq) if seq else len(qa_list) + 1,
            "question": question.strip(),
            "answer": answer.strip(),
        })

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(qa_list, f, ensure_ascii=False, indent=2)

    print(f"预处理完成：共 {len(qa_list)} 条有效问答对")
    print(f"输出文件：{os.path.abspath(OUTPUT_PATH)}")


if __name__ == "__main__":
    preprocess()
